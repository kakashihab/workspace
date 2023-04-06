import openpyxl
from openpyxl.utils import column_index_from_string

workbook = openpyxl.load_workbook("Week 2\gdp.xlsx")

earnings_sheet = workbook["Metadata - Countries"]
countries_sheet = workbook["Data"]
countries = earnings_sheet.iter_rows(min_col = 1, max_col = 3, min_row = 2, values_only = True)
gdp = countries_sheet.iter_rows(min_col = 1, min_row = 5, values_only = True)

year_2000 = column_index_from_string("AS")
year_2010 = column_index_from_string("BC")
year_2019 = column_index_from_string("BL")

high_earners = []
years = {"2000" : year_2000, "2010" : year_2010, "2019": year_2019}

for country in countries :
    if country[2] == 'High income':
        high_earners.append(country[0])

for country in gdp :
    if country[1] in high_earners :
        for year in years :
            if country[years[year]] :
                print(f"The GDP of {country[0]} in the year {year} is ${country[years[year]]}")

workbook.close()