import openpyxl

# open a xlsx file (workbook)
wb = openpyxl.load_workbook("Week 2\gdp.xlsx")
wb.create_sheet("LA&C")

new_sheet = wb["LA&C"]
new_sheet["A1"] = "Latin America and Caribbean GDP"
new_sheet["A3"] = "Country code"
new_sheet["B3"] = "Country name"
new_sheet["C3"] = "1990"
new_sheet["D3"] = "2019"


# get the names of the sheets in the workbook
#print(wb.sheetnames)

# open a specific sheet
ws = wb["Data"]
wsm = wb["Metadata - Countries"]

country_regions = wsm.iter_rows(min_col = 1, max_col = 2, min_row = 2, values_only = True)
for country_region in country_regions:
    print("{} : {}".format(country_region[0], country_region[1]))

    dict_regions = dict(country_regions)
    


country_names = ws.iter_rows(min_col = 1, max_col = 64, min_row = 5, values_only = True)
for country_name in country_names:
    if (country_name[1] in dict_regions) and (dict_regions[country_name[1]] == "Latin America & Caribbean"):
        region_code = dict_regions[country_name[1]]
        print("{} : {} : {} : {}".format(country_name[0], region_code, country_name[34], country_name[63]))
        new_sheet.append((country_name[1], country_name[0], country_name[34], country_name[63]))

wb.save("Week 2\gdp_modified.xlsx")
# close the file when we are finished with it
wb.close()